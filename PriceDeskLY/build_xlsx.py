import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference

T = json.load(open('data/tasks_final.json'))
DRV = {'fixed':'Fixed','site':'Per site','visit':'Per visit','month':'Per month',
       'sitemonth':'Per site-month','week':'Per week','mvisit':'Per monitoring visit',
       'patient':'Per patient','manual':'Manual'}
DEPT_FULL = {'Medical':'Medical writing','SSU':'Site start-up','DM':'Data management',
  'PM':'Project management','SM':'Site monitoring','EC':'Regulatory / EC','BioStats':'Biostatistics',
  'PTC':'Pass-through costs','Tech':'Technology licences'}
DEPTS = ['Medical','SSU','DM','PM','SM','EC','BioStats','PTC','Tech']

F   = 'Arial'
INR = '[>=10000000]#\\,##\\,##\\,##0;[>=100000]#\\,##\\,##0;#,##0'
PCT = '0.0%'
NUM = '#,##0.##'
BLUE  = Font(name=F, size=10, color='0000FF')          # user input
BLACK = Font(name=F, size=10)                          # formula
GREEN = Font(name=F, size=10, color='008000')          # link to another sheet
BOLD  = Font(name=F, size=10, bold=True)
H1    = Font(name=F, size=14, bold=True, color='0B4F49')
HDR   = Font(name=F, size=9, bold=True, color='FFFFFF')
HDRFILL = PatternFill('solid', fgColor='0B4F49')
BANDFILL= PatternFill('solid', fgColor='EEF3F2')
YELLOW  = PatternFill('solid', fgColor='FFFF00')
NOTE  = Font(name=F, size=8, italic=True, color='606060')
thin  = Side(style='thin', color='C8D4D2')
BOX   = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook(); wb.remove(wb.active)

# ───────────────────────────── Inputs ─────────────────────────────
ws = wb.create_sheet('Inputs')
ws['B2']='PriceDeskLY'; ws['B2'].font=H1
ws['B3']='Study pricing model - all amounts in INR'; ws['B3'].font=NOTE
ws['B5']='HOW TO USE'; ws['B5'].font=BOLD
for i,t in enumerate([
  'Blue cells are yours to edit. Black cells are formulas - do not overwrite them.',
  'Set the study drivers and commercial levers below, then mark lines with "x" in column A of the Quote sheet.',
  'Margin is applied line by line, then summed. It is a share of the invoice, not a mark-up on cost.',
  'Pass-through and Tech lines bill at cost: no margin, no discount, excluded from revenue.',
  'Summary shows the roll-up. Checks must all read OK before the quote goes out.']):
    ws.cell(6+i,2,'- '+t).font=NOTE

r=13
ws.cell(r-1,2,'STUDY DRIVERS').font=BOLD
drivers=[('Sites',5),('Patients enrolled',250),('Visits per patient',3),
         ('Study duration (months)',12),('Monitoring visits per site',4)]
names={}
for i,(lab,val) in enumerate(drivers):
    ws.cell(r+i,2,lab).font=BLACK
    c=ws.cell(r+i,3,val); c.font=BLUE; c.number_format='#,##0'; c.border=BOX; c.fill=YELLOW
names.update({'Sites':'C13','Patients':'C14','VisitsPerPatient':'C15','Months':'C16','MonVisitsPerSite':'C17'})

r=20
ws.cell(r-1,2,'DERIVED').font=BOLD
derived=[('Total patient visits','=Patients*VisitsPerPatient','TotalVisits'),
         ('Total monitoring visits','=Sites*MonVisitsPerSite','MonVisits'),
         ('Site-months','=Sites*Months','SiteMonths'),
         ('Weeks','=ROUND(Months*4.345,0)','Weeks')]
for i,(lab,fml,nm) in enumerate(derived):
    ws.cell(r+i,2,lab).font=BLACK
    c=ws.cell(r+i,3,fml); c.font=BLACK; c.number_format='#,##0'; c.border=BOX
    names[nm]=f'C{r+i}'

r=26
ws.cell(r-1,2,'COMMERCIAL LEVERS').font=BOLD
ws.cell(r,2,'Margin on invoice').font=BLACK
c=ws.cell(r,3,0.225); c.font=BLUE; c.number_format=PCT; c.border=BOX; c.fill=YELLOW
ws.cell(r,4,'share of the invoice, not a mark-up on cost').font=NOTE
ws.cell(r+1,2,'Margin floor (policy)').font=BLACK
c=ws.cell(r+1,3,0.225); c.font=BLUE; c.number_format=PCT; c.border=BOX
ws.cell(r+1,4,'22.5% minimum on every sale').font=NOTE
ws.cell(r+2,2,'Client discount').font=BLACK
c=ws.cell(r+2,3,0.0); c.font=BLUE; c.number_format=PCT; c.border=BOX; c.fill=YELLOW
ws.cell(r+2,4,'applies to professional fees only, never to pass-through').font=NOTE
ws.cell(r+3,2,'Margin needed for this discount').font=BLACK
c=ws.cell(r+3,3,'=1-(1-MarginFloor)*(1-Discount)'); c.font=BLACK; c.number_format=PCT; c.border=BOX
ws.cell(r+3,4,'set Margin on invoice to at least this to hold the floor').font=NOTE
names.update({'MarginLever':'C26','MarginFloor':'C27','Discount':'C28'})

r=33
ws.cell(r-1,2,'DRIVER TO QUANTITY').font=BOLD
ws.cell(r-1,4,'the Quote sheet looks quantities up here').font=NOTE
dmap=[('Fixed','=1'),('Per site','=Sites'),('Per visit','=TotalVisits'),('Per month','=Months'),
      ('Per site-month','=SiteMonths'),('Per week','=Weeks'),('Per monitoring visit','=MonVisits'),
      ('Per patient','=Patients'),('Manual','=1')]
ws.cell(r,2,'Driver').font=HDR; ws.cell(r,2).fill=HDRFILL
ws.cell(r,3,'Quantity').font=HDR; ws.cell(r,3).fill=HDRFILL
for i,(lab,fml) in enumerate(dmap):
    ws.cell(r+1+i,2,lab).font=BLACK; ws.cell(r+1+i,2).border=BOX
    c=ws.cell(r+1+i,3,fml); c.font=BLACK; c.number_format='#,##0'; c.border=BOX
names['DriverName']=f'$B${r+1}:$B${r+len(dmap)}'
names['DriverQty'] =f'$C${r+1}:$C${r+len(dmap)}'
ws.column_dimensions['A'].width=2; ws.column_dimensions['B'].width=32
ws.column_dimensions['C'].width=14; ws.column_dimensions['D'].width=52
ws.sheet_view.showGridLines=False

for nm,ref in names.items():
    ref = f"Inputs!{ref}" if ref.startswith('$') else f"Inputs!${ref[0]}${ref[1:]}"
    wb.defined_names.add(DefinedName(nm, attr_text=ref))

# ───────────────────────────── Rate Card ─────────────────────────────
rc = wb.create_sheet('Rate Card')
hdr=['Department','Task','Kind','Driver','Delivery hrs','Delivery rate','Review hrs','Review rate','Unit cost','Rate set?']
for j,h in enumerate(hdr,1):
    c=rc.cell(1,j,h); c.font=HDR; c.fill=HDRFILL; c.border=BOX
    c.alignment=Alignment(horizontal='right' if j>=5 else 'left', wrap_text=True)
for i,t in enumerate(T):
    r=i+2; svc = t['kind']=='hours'
    rc.cell(r,1,t['dept']).font=BLACK
    rc.cell(r,2,t['task']).font=BLACK
    rc.cell(r,3,'Service' if svc else 'Pass-through').font=BLACK
    rc.cell(r,4,DRV[t['driver']]).font=BLACK
    if svc:
        for col,val in ((5,t['fte']),(6,t['frate']),(7,t['rev']),(8,t['rrate'])):
            c=rc.cell(r,col, val if val is not None else None)
            c.font=BLUE; c.number_format=NUM if col in (5,7) else INR
        c=rc.cell(r,9,f'=IFERROR(N(E{r})*N(F{r})+N(G{r})*N(H{r}),0)'); c.font=BLACK
    else:
        c=rc.cell(r,9, t['unit'] if t['unit'] else None); c.font=BLUE
    c.number_format=INR
    f=rc.cell(r,10,f'=IF(N(I{r})>0,"OK","RATE NOT SET")'); f.font=BLACK
    if i%2: 
        for j in range(1,11): rc.cell(r,j).fill=BANDFILL
NROW=len(T)+1
rc.auto_filter.ref=f'A1:J{NROW}'
rc.freeze_panes='C2'
for col,w in zip('ABCDEFGHIJ',[12,62,13,20,12,13,12,12,14,13]): rc.column_dimensions[col].width=w
rc.conditional_formatting.add(f'J2:J{NROW}',
    CellIsRule(operator='equal', formula=['"RATE NOT SET"'],
               font=Font(name=F,size=10,bold=True,color='B3261E'),
               fill=PatternFill('solid',fgColor='FCECE9')))
rc.cell(NROW+2,2,'Delivery rate = department hourly cost. Review rate = INR 3,000/hr senior review.').font=NOTE
rc.cell(NROW+3,2,'Source: client rate card supplied 20 Aug 2026. Blank hours mean the rate was never set - fill them in.').font=NOTE

# ───────────────────────────── Quote ─────────────────────────────
q = wb.create_sheet('Quote')
qh=['In','Department','Activity','Kind','Driver','Qty','Unit cost','Margin override',
    'Margin applied','Unit price','Line price','Line cost','Delivery hrs','Review hrs']
for j,h in enumerate(qh,1):
    c=q.cell(1,j,h); c.font=HDR; c.fill=HDRFILL; c.border=BOX
    c.alignment=Alignment(horizontal='center' if j==1 else ('right' if j>=6 else 'left'), wrap_text=True)
DEFAULT_ON = {t['task'] for t in T[:0]}
for i,t in enumerate(T):
    r=i+2; rcr=i+2; svc=t['kind']=='hours'
    c=q.cell(r,1,None); c.font=BLUE; c.alignment=Alignment(horizontal='center'); c.border=BOX; c.fill=YELLOW
    q.cell(r,2,f"='Rate Card'!A{rcr}").font=GREEN
    q.cell(r,3,f"='Rate Card'!B{rcr}").font=GREEN
    q.cell(r,4,f"='Rate Card'!C{rcr}").font=GREEN
    q.cell(r,5,f"='Rate Card'!D{rcr}").font=GREEN
    c=q.cell(r,6,f'=IF($A{r}="x",IFERROR(INDEX(DriverQty,MATCH($E{r},DriverName,0)),0),0)')
    c.font=BLACK; c.number_format='#,##0.##'
    c=q.cell(r,7,f"='Rate Card'!I{rcr}"); c.font=GREEN; c.number_format=INR
    c=q.cell(r,8,None); c.font=BLUE; c.number_format=PCT; c.border=BOX
    c=q.cell(r,9,f'=IF($D{r}="Pass-through",0,MAX(MarginFloor,IF($H{r}="",MarginLever,$H{r})))')
    c.font=BLACK; c.number_format=PCT
    c=q.cell(r,10,f'=IF($D{r}="Pass-through",$G{r},IF($I{r}>=1,$G{r},ROUND($G{r}/(1-$I{r}),0)))')
    c.font=BLACK; c.number_format=INR
    c=q.cell(r,11,f'=$F{r}*$J{r}'); c.font=BLACK; c.number_format=INR
    c=q.cell(r,12,f'=$F{r}*$G{r}'); c.font=BLACK; c.number_format=INR
    c=q.cell(r,13,f"=$F{r}*N('Rate Card'!E{rcr})"); c.font=BLACK; c.number_format=NUM
    c=q.cell(r,14,f"=$F{r}*N('Rate Card'!G{rcr})"); c.font=BLACK; c.number_format=NUM
    if i%2:
        for j in range(2,15):
            if j not in (1,8): q.cell(r,j).fill=BANDFILL
dv=DataValidation(type='list', formula1='"x"', allow_blank=True,
                  prompt='Type x to include this line in the quote', promptTitle='Include')
q.add_data_validation(dv); dv.add(f'A2:A{NROW}')
q.auto_filter.ref=f'A1:N{NROW}'
q.freeze_panes='D2'
for col,w in zip('ABCDEFGHIJKLMN',[5,12,58,13,20,10,13,12,12,13,14,14,11,11]):
    q.column_dimensions[col].width=w
q.conditional_formatting.add(f'A2:N{NROW}',
    FormulaRule(formula=[f'$A2="x"'], fill=PatternFill('solid',fgColor='E3F0ED')))
q.conditional_formatting.add(f'G2:G{NROW}',
    FormulaRule(formula=[f'AND($A2="x",$G2=0)'],
        font=Font(name=F,size=10,bold=True,color='B3261E'), fill=PatternFill('solid',fgColor='FCECE9')))
q.conditional_formatting.add(f'H2:H{NROW}',
    FormulaRule(formula=[f'AND($H2<>"",$H2<MarginFloor)'],
        font=Font(name=F,size=10,bold=True,color='B3261E'), fill=PatternFill('solid',fgColor='FCECE9')))
q.cell(NROW+2,3,'Mark a line with "x" in column A to include it. Margin override is optional - leave blank to use the lever on Inputs.').font=NOTE
q.cell(NROW+3,3,'Unit price = unit cost / (1 - margin), rounded to whole rupees, so Qty x Unit price equals Line price exactly.').font=NOTE

wb.save('PriceDeskLY_Model.xlsx')
print('written; rows =',NROW)
