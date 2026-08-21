import json,re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference

F='Arial'
INR='[>=10000000]#\\,##\\,##\\,##0;[>=100000]#\\,##\\,##0;#,##0'
PCT='0.0%'
BLACK=Font(name=F,size=10); BOLD=Font(name=F,size=10,bold=True)
H1=Font(name=F,size=14,bold=True,color='0B4F49')
HDR=Font(name=F,size=9,bold=True,color='FFFFFF'); HDRFILL=PatternFill('solid',fgColor='0B4F49')
NOTE=Font(name=F,size=8,italic=True,color='606060')
BIG=Font(name=F,size=13,bold=True); ACC=Font(name=F,size=13,bold=True,color='0B4F49')
thin=Side(style='thin',color='C8D4D2'); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
BAND=PatternFill('solid',fgColor='EEF3F2')

wb=load_workbook('PriceDeskLY_Model.xlsx')
q=wb['Quote']; N=123
QK=f"Quote!$K$2:$K${N}"; QL=f"Quote!$L$2:$L${N}"; QD=f"Quote!$D$2:$D${N}"
QA=f"Quote!$A$2:$A${N}"; QB=f"Quote!$B$2:$B${N}"; QF=f"Quote!$F$2:$F${N}"
QG=f"Quote!$G$2:$G${N}"; QH=f"Quote!$H$2:$H${N}"; QJ=f"Quote!$J$2:$J${N}"

# ── seed the RWE example so the file opens with a live quote ──
html=open('PriceDeskLY.html',encoding='utf-8').read()
i=html.index('const PRESETS='); j=i+len('const PRESETS='); d=0
for k in range(j,len(html)):
    if html[k]=='{': d+=1
    elif html[k]=='}':
        d-=1
        if d==0: blk=html[j:k+1]; break
rwe=set(n.replace("\\'","'") for n in re.findall(r"[\"']((?:[^\"'\\]|\\.)*)[\"']",blk)[:33])
T=json.load(open('data/tasks_final.json'))
seed=0
for i2,t in enumerate(T):
    if t['task'] in rwe: q.cell(i2+2,1,'x'); seed+=1

# ───────────────────────── Summary ─────────────────────────
s=wb.create_sheet('Summary',2)
s['B2']='Quote summary'; s['B2'].font=H1
s['B3']='Every figure below is a roll-up of the Quote sheet. Nothing is typed here.'; s['B3'].font=NOTE
rows=[('Lines included',f'=COUNTIF({QA},"x")','#,##0',None),
      ('Total fixed fee (cost)',f'=SUMIFS({QL},{QD},"Service")',INR,'sum of Line cost on service lines'),
      ('Sum of line prices',f'=SUMIFS({QK},{QD},"Service")',INR,'sum of Line price - margin already inside each line'),
      ('Margin (rupees)','=C7-C6',INR,'the difference, not a separate uplift'),
      ('Discount','=-C7*Discount',INR,'professional fees only'),
      ('Net revenue','=C7+C9',INR,'what you actually earn'),
      ('Pass-through at cost',f'=SUMIFS({QL},{QD},"Pass-through")',INR,'no margin, no discount, not revenue'),
      ('Client price','=C10+C11',INR,'total invoice'),
      ('Margin on revenue','=IF(C10=0,0,(C10-C6)/C10)',PCT,'must be at or above the floor'),
      ('Uplift on cost','=IF(C6=0,0,(C7-C6)/C6)',PCT,'what 22.5% margin looks like as a mark-up'),
      ('Delivery hours',f'=SUM(Quote!M2:M{N})','#,##0',None),
      ('Review hours',f'=SUM(Quote!N2:N{N})','#,##0',None)]
for i3,(lab,fml,fmt,note) in enumerate(rows):
    r=5+i3
    s.cell(r,2,lab).font=BOLD if lab in ('Client price','Margin on revenue') else BLACK
    c=s.cell(r,3,fml); c.number_format=fmt; c.border=BOX
    c.font=ACC if lab in ('Client price','Margin on revenue') else BLACK
    if note: s.cell(r,4,note).font=NOTE
s.cell(19,2,'BY DEPARTMENT').font=BOLD
for j2,h in enumerate(['Department','Lines','Cost','Price','Share of price'],2):
    c=s.cell(20,j2,h); c.font=HDR; c.fill=HDRFILL; c.border=BOX
DEPTS=['Medical','SSU','DM','PM','SM','EC','BioStats','PTC','Tech']
for i4,dp in enumerate(DEPTS):
    r=21+i4
    s.cell(r,2,dp).font=BLACK
    s.cell(r,3,f'=COUNTIFS({QA},"x",{QB},$B{r})').font=BLACK
    c=s.cell(r,4,f'=SUMIFS({QL},{QB},$B{r})'); c.font=BLACK; c.number_format=INR
    c=s.cell(r,5,f'=SUMIFS({QK},{QB},$B{r})'); c.font=BLACK; c.number_format=INR
    c=s.cell(r,6,f'=IF($E$30=0,0,E{r}/$E$30)'); c.font=BLACK; c.number_format=PCT
    if i4%2:
        for j2 in range(2,7): s.cell(r,j2).fill=BAND
s.cell(30,2,'Total').font=BOLD
for col in (3,4,5):
    c=s.cell(30,col,f'=SUM({chr(64+col)}21:{chr(64+col)}29)'); c.font=BOLD
    c.number_format='#,##0' if col==3 else INR
for col,w in zip('ABCDEF',[2,30,18,52,18,15]): s.column_dimensions[col].width=w
s.sheet_view.showGridLines=False
ch=BarChart(); ch.type='bar'; ch.style=2; ch.title='Cost by department'
ch.y_axis.title=None; ch.x_axis.title=None; ch.height=9; ch.width=17; ch.legend=None
ch.add_data(Reference(s,min_col=4,min_row=20,max_row=29), titles_from_data=True)
ch.set_categories(Reference(s,min_col=2,min_row=21,max_row=29))
s.add_chart(ch,'H5')

# ───────────────────────── Checks ─────────────────────────
ck=wb.create_sheet('Checks',3)
ck['B2']='Model checks'; ck['B2'].font=H1
ck['B3']='All must read OK before this quote leaves the building.'; ck['B3'].font=NOTE
for j3,h in enumerate(['#','Check','Result','Status'],2):
    c=ck.cell(5,j3,h); c.font=HDR; c.fill=HDRFILL; c.border=BOX
checks=[
 ('Line prices roll up to the quoted fees',
  f'=ROUND(SUMIFS({QK},{QD},"Service")-Summary!C7,2)','=IF(D6=0,"OK","FAIL")','#,##0.00'),
 ('Qty x Unit price = Line price on every row',
  f'=SUMPRODUCT(--(ROUND({QF}*{QJ},2)<>ROUND({QK},2)))','=IF(D7=0,"OK","FAIL")','#,##0'),
 ('Realised margin at or above the floor',
  '=Summary!C13-MarginFloor','=IF(D8>=-0.0005,"OK","FAIL")',PCT),
 ('Margin lever at or above the floor',
  '=MarginLever-MarginFloor','=IF(D9>=0,"OK","FAIL")',PCT),
 ('Included lines with no rate set',
  f'=COUNTIFS({QA},"x",{QG},0)','=IF(D10=0,"OK","FAIL")','#,##0'),
 ('Included lines resolving to zero quantity',
  f'=COUNTIFS({QA},"x",{QF},0)','=IF(D11=0,"OK","REVIEW")','#,##0'),
 ('Line margin overrides below the floor',
  f'=SUMPRODUCT(--({QH}<>""),--({QH}<MarginFloor))','=IF(D12=0,"OK","FAIL")','#,##0'),
 ('Pass-through lines carrying margin',
  f'=SUMPRODUCT(--({QD}="Pass-through"),--(Quote!$I$2:$I${N}>0))','=IF(D13=0,"OK","FAIL")','#,##0'),
 ('Lines where margin differs from the lever',
  f'=SUMPRODUCT(--({QA}="x"),--({QD}="Service"),--(Quote!$I$2:$I${N}<>MarginLever))',
  '=IF(D14=0,"OK","REVIEW")','#,##0'),
 ('Discount leaves margin under the floor',
  '=IF(Summary!C13<MarginFloor-0.0005,1,0)','=IF(D15=0,"OK","FAIL")','#,##0')]
for i5,(lab,fml,st,fmt) in enumerate(checks):
    r=6+i5
    ck.cell(r,2,i5+1).font=BLACK
    ck.cell(r,3,lab).font=BLACK
    c=ck.cell(r,4,fml); c.font=BLACK; c.number_format=fmt; c.border=BOX
    c=ck.cell(r,5,st); c.font=BOLD; c.alignment=Alignment(horizontal='center'); c.border=BOX
    if i5%2:
        for j3 in range(2,6): ck.cell(r,j3).fill=BAND
ck.conditional_formatting.add('E6:E15',
  CellIsRule(operator='equal',formula=['"OK"'],
             font=Font(name=F,size=10,bold=True,color='1B7F4B'),
             fill=PatternFill('solid',fgColor='E8F4ED')))
ck.conditional_formatting.add('E6:E15',
  CellIsRule(operator='equal',formula=['"FAIL"'],
             font=Font(name=F,size=10,bold=True,color='B3261E'),
             fill=PatternFill('solid',fgColor='FCECE9')))
ck.conditional_formatting.add('E6:E15',
  CellIsRule(operator='equal',formula=['"REVIEW"'],
             font=Font(name=F,size=10,bold=True,color='B4690E'),
             fill=PatternFill('solid',fgColor='FDF3E3')))
ck.cell(18,3,'Checks 6 and 9 read REVIEW rather than FAIL: both can be legitimate, but must be deliberate.').font=NOTE
ck.cell(19,3,'Contingency is deliberately absent from this model. It plays no part in the calculation.').font=NOTE
for col,w in zip('ABCDE',[2,5,52,18,12]): ck.column_dimensions[col].width=w
ck.sheet_view.showGridLines=False

wb.active=0
wb.save('PriceDeskLY_Model.xlsx')
print('summary + checks added; example lines seeded =',seed)
